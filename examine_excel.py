#!/usr/bin/env python3
"""
Script to examine the structure of predictions.xlsx file
"""
import openpyxl
import sys
from pathlib import Path

def examine_excel(file_path):
    """Examine Excel file structure and print detailed information"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"📊 Excel File Analysis: {file_path}")
    print(f"{'='*80}\n")
    
    print(f"📄 Worksheet Name: {ws.title}")
    print(f"📐 Dimensions: {ws.dimensions}")
    print(f"📏 Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
    
    # Print headers (row 1)
    print("📋 COLUMN HEADERS (Row 1):")
    print("-" * 80)
    headers = []
    for col in range(1, min(ws.max_column + 1, 50)):  # Limit to first 50 columns
        cell_value = ws.cell(1, col).value
        headers.append(cell_value)
        print(f"Column {col:2d}: {cell_value}")
    
    # Print first 5 data rows
    print(f"\n📊 SAMPLE DATA (Rows 2-6):")
    print("-" * 80)
    for row in range(2, min(ws.max_row + 1, 7)):
        print(f"\n🔹 Row {row}:")
        for col in range(1, min(ws.max_column + 1, 30)):  # Show first 30 columns
            cell_value = ws.cell(row, col).value
            header = headers[col-1] if col-1 < len(headers) else f"Col{col}"
            if cell_value is not None and str(cell_value).strip():
                print(f"  [{col:2d}] {header}: {cell_value}")
    
    # Analyze which columns the ExtractFromExcel.cs uses
    print(f"\n🔍 COLUMNS USED BY ExtractFromExcel.cs:")
    print("-" * 80)
    column_mapping = {
        2: "HomeTeam",
        3: "AwayTeam",
        4: "League",
        5: "Date/Time",
        6: "HomeWin",
        7: "Draw",
        8: "AwayWin",
        18: "OverTwoGoals",
        22: "OverThreeGoals",
        24: "OverFourGoals",
        34: "UnderTwoGoals",
        38: "UnderThreeGoals"
    }
    
    print("\nMapping from C# code (ExtractFromExcel.cs):")
    for col_num, field_name in sorted(column_mapping.items()):
        header = headers[col_num-1] if col_num-1 < len(headers) else "N/A"
        sample_val = ws.cell(2, col_num).value if ws.max_row >= 2 else "N/A"
        print(f"  Col {col_num:2d} → {field_name:20s} | Header: '{header}' | Sample: '{sample_val}'")
    
    # Statistics
    print(f"\n📈 DATA STATISTICS:")
    print("-" * 80)
    total_rows = ws.max_row - 1  # Exclude header
    print(f"Total data rows: {total_rows}")
    
    # Count matches for today (as the code filters by current day)
    today_count = 0
    from datetime import datetime
    current_day = datetime.now().day
    
    for row in range(2, ws.max_row + 1):
        date_str = ws.cell(row, 5).value
        if date_str:
            date_str = str(date_str)
            try:
                day = int(date_str.split('.')[0].split(' ')[0])
                if day == current_day:
                    today_count += 1
            except:
                pass
    
    print(f"Matches for today (day {current_day}): {today_count}")
    print(f"Total matches in file: {total_rows}")

if __name__ == "__main__":
    file_path = Path(__file__).parent / "Resources" / "predictions.xlsx"
    
    if not file_path.exists():
        print(f"❌ Error: File not found at {file_path}")
        print("Looking for alternative paths...")
        alt_paths = [
            Path(__file__).parent / "MatchPredictor.Web" / "Resources" / "predictions.xlsx",
            Path("Resources") / "predictions.xlsx",
        ]
        for alt_path in alt_paths:
            if alt_path.exists():
                file_path = alt_path
                print(f"✅ Found file at: {file_path}")
                break
        else:
            sys.exit(1)
    
    try:
        examine_excel(file_path)
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
