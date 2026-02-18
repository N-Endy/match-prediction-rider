#!/usr/bin/env python3
"""Check all worksheets in predictions.xlsx"""
import openpyxl

wb = openpyxl.load_workbook("Resources/predictions.xlsx")
print(f"Total worksheets: {len(wb.worksheets)}\n")

for idx, ws in enumerate(wb.worksheets):
    print(f"Worksheet {idx + 1}: '{ws.title}'")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Print first row headers for each sheet
    print(f"  Headers: ", end="")
    headers = []
    for col in range(1, min(ws.max_column + 1, 50)):
        val = ws.cell(1, col).value
        if val:
            headers.append(str(val))
    print(", ".join(headers))
    print()
