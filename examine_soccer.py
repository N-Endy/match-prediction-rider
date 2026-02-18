#!/usr/bin/env python3
"""Examine the soccer worksheet in detail"""
import openpyxl

wb = openpyxl.load_workbook("Resources/predictions.xlsx")
ws = wb["soccer"]  # Get soccer sheet specifically

print("⚽ SOCCER WORKSHEET ANALYSIS")
print("=" * 100)
print(f"Dimensions: {ws.dimensions}")
print(f"Total rows: {ws.max_row}, Total columns: {ws.max_column}\n")

# Print all headers with column numbers
print("📋 ALL COLUMN HEADERS:")
print("-" * 100)
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    headers.append(val)
    print(f"Col {col:2d}: {val}")

# Mapping from ExtractFromExcel.cs
print("\n🔍 COLUMNS USED BY C# CODE (ExtractFromExcel.cs):")
print("-" * 100)
column_mapping = {
    2: "HomeTeam",
    3: "AwayTeam",
    4: "League",
    5: "Date/Time",
    6: "HomeWin (1x2_h)",
    7: "Draw (1x2_d)",
    8: "AwayWin (1x2_a)",
    18: "OverTwoGoals (o_2.5)",
    22: "OverThreeGoals (o_3)",
    24: "OverFourGoals (o_4)",
    34: "UnderTwoGoals (u_2.5)",
    38: "UnderThreeGoals (u_3)"
}

for col_num, field_description in sorted(column_mapping.items()):
    header = headers[col_num-1] if col_num-1 < len(headers) else "OUT OF RANGE"
    sample_val = ws.cell(2, col_num).value if ws.max_row >= 2 else "N/A"
    print(f"  Col {col_num:2d} → {field_description:30s} | Header: '{header}' | Sample: '{sample_val}'")

# Print sample rows
print("\n📊 SAMPLE SOCCER MATCHES (First 3 rows):")
print("-" * 100)
for row in range(2, min(ws.max_row + 1, 5)):
    print(f"\n🔹 Row {row}:")
    # Print key columns only
    key_cols = [1, 2, 3, 4, 5, 6, 7, 8, 18, 22, 24, 34, 38]
    for col in key_cols:
        if col <= ws.max_column:
            val = ws.cell(row, col).value
            header = headers[col-1] if col-1 < len(headers) else f"Col{col}"
            print(f"  [{col:2d}] {header}: {val}")

# Count today's matches
from datetime import datetime
current_day = datetime.now().day
today_count = 0

for row in range(2, ws.max_row + 1):
    date_str = ws.cell(row, 5).value
    if date_str:
        date_str = str(date_str)
        try:
            day = int(date_str.split('.')[0].split(' ')[0])
            if day == current_day:
                today_count += 1
        except:
            pass

print(f"\n📈 STATISTICS:")
print("-" * 100)
print(f"Total soccer matches in file: {ws.max_row - 1}")
print(f"Matches for today (day {current_day}): {today_count}")

# Analyze data quality
print(f"\n🔬 DATA QUALITY ANALYSIS:")
print("-" * 100)

# Check for null values in key columns
null_counts = {}
for col_num in [6, 7, 8, 18, 22, 34]:
    null_count = sum(1 for row in range(2, ws.max_row +1) if ws.cell(row, col_num).value is None)
    col_name = headers[col_num-1] if col_num-1 < len(headers) else f"Col{col_num}"
    null_counts[col_name] = null_count
    print(f"  {col_name}: {null_count} null values ({null_count/(ws.max_row-1)*100:.1f}%)")

print(f"\n✅ CONCLUSION:")
print("-" * 100)
print(f"The Excel file contains pre-calculated PROBABILITIES from sports-ai.dev")
print(f"NOT historical match scores - these are predictions for upcoming matches!")
print(f"The system scrapes these AI-generated probabilities and applies additional logic.")
